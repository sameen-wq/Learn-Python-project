import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook
wb = xl.load_workbook('transactions3.xlsx')
sheet = wb['Sheet1']

# Calculate discount and total price for each row
for row in range(2, sheet.max_row + 1):
    cell=sheet.cell(row, 3)

    result=cell.value * 0.2
    discount_price = sheet.cell(row, 4)
    discount_price.value=result

    product_price=sheet.cell(row, 3)
    price_to_pay=(product_price.value - result)
    table_price_to_pay=sheet.cell(row, 5)
    table_price_to_pay.value=price_to_pay

values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=5,
          max_col=5)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'F5')

wb.save('transaction33.xlsx')
