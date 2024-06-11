import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
# Load the workbook and select the sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Process each row and calculate corrected prices
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)  # Assuming column C has the original prices
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)  # Write the corrected price to column D
        corrected_price_cell.value = corrected_price

    # Create a reference to the corrected prices in column D
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    # Create and add the bar chart
    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    sheet.add_chart(chart, 'E2')

    # Save the workbook with a new name
    wb.save(filename)
